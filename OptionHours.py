import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
	options =  pickle.load(file)

wb = load_workbook(r"templates\HoursTemplate.xlsx") # create new workbook
for length in ["18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36"]:
	ws = wb[length]
	bold = Font(bold=True, underline="single")
	red = Font(color="FF0000")
	blue = Font(color="0000FF")
	#ws = wb.active # selcet the active worksheet
	row = 1
	for option in sorted(options): #18
		row += 1
		ws.cell(row = row, column = 1).value = option
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value = options[option][length + " DESIGN HOURS"]
		ws.cell(row = row, column = 3).value = options[option]["FABRICATION " + length + " HOURS"]
		ws.cell(row = row, column = 4).value = options[option]["CANVAS " + length + " HOURS"]
		ws.cell(row = row, column = 5).value = options[option]["PAINT " + length + " HOURS"]
		ws.cell(row = row, column = 6).value = options[option]["OUTFITTING " + length + " HOURS"]
			
wb.save(r"output\Option Hours.xlsx")

