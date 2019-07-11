import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
	options =  pickle.load(file)

kill=[]
for option in options:
	co = options.get(option+" - CO")
	cr = options.get(option+" - CR")
	if co:
		kill.append(option+" - CO")
		options[option]["OUTFITTING PARTS"] += co["OUTFITTING PARTS"]
		options[option]["CANVAS PARTS"] += co["CANVAS PARTS"]
		options[option]["PAINT PARTS"] += co["PAINT PARTS"]
		options[option]["FABRICATION PARTS"] += co["FABRICATION PARTS"]
	if cr:
		kill.append(option+" - CR")
		options[option]["OUTFITTING PARTS"] += cr["OUTFITTING PARTS"]
		options[option]["CANVAS PARTS"] += cr["CANVAS PARTS"]
		options[option]["PAINT PARTS"] += cr["PAINT PARTS"]
		options[option]["FABRICATION PARTS"] += cr["FABRICATION PARTS"]

for option in kill:
	del options[option]
		
wb = load_workbook("templates\CostingSheetTemplate.xlsx") # create new workbook
ws = wb['L18']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #18

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["18 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
			
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["18 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["18 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L19']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #19

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["19 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["19 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["19 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L20']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #20

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["20 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["20 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["20 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L21']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #21

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["21 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["21 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["21 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L22']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #22

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["22 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["22 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["22 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L23']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #23

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["23 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["23 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["23 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L24']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #24

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["24 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["24 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["24 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L25']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #25

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["25 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["25 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["25 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L26']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #26

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["26 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["26 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["26 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L27']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #27

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["27 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["27 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["27 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L28']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #28

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["28 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["28 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["28 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L29']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #29

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["29 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["29 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["29 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L30']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #30

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["30 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["30 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["30 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L31']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #31

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["31 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["31 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["31 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L32']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #32

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["32 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["32 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["32 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")


wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L33']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #33

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["33 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["33 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["33 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L34']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #34

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["34 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["34 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["34 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L35']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #35

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["35 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["35 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["35 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L36']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options): #36

	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

			row += 1
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 1).font = bold
			ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
			ws.cell(row = row, column = 2).font = bold
	if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"]
		for item in rigging:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Outfitting"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["36 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Fabrication"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["36 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'
	
	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
		
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = "Paint"
			ws.cell(row = row, column = 3).value = item["VENDOR"]
			ws.cell(row = row, column = 4).value = item["VENDOR PART"]
			ws.cell(row = row, column = 5).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 6).value = float(item["PRICE"])
			ws.cell(row = row, column = 6).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 7).value = item["UOM"]
			ws.cell(row = row, column = 8).value = item["36 QTY"]
			ws.cell(row = row, column = 9).value = "=SUM(F" + str(row) + "*H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = 0
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 11).value = "=SUM(H" + str(row) + "+I" + str(row) + ")"
			ws.cell(row = row, column = 11).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")