import pickle
from openpyxl import load_workbook


# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
    options =  pickle.load(file)

wb = load_workbook("templates\CostingSheetTemplate.xlsx") # create new workbook
ws = wb['L18']
#ws = wb.active # selcet the active worksheet

for row, option in enumerate(sorted(options), start = 2):
	#for options:
	for item in options[option]["OUTFITTING PARTS"]:
		if item ["PART NUMBER"]:
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 3).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["PRICE"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
	#for options:
	for item in options[option]["CANVAS PARTS"]:
		if item ["PART NUMBER"]:
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 3).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["PRICE"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L19']
#ws = wb.activ # selcet the active worksheet

for row, option in enumerate(sorted(options), start = 2):
	#for options:
	for item in options[option]["OUTFITTING PARTS"]:
		if item ["PART NUMBER"]:
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 3).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["PRICE"]
			ws.cell(row = row, column = 7).value = item["19 QTY"]
	#for options:
	for item in options[option]["CANVAS PARTS"]:
		if item ["PART NUMBER"]:
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 3).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["PRICE"]
			ws.cell(row = row, column = 7).value = item["19 QTY"]
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")