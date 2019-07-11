import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
    options =  pickle.load(file)

wb = load_workbook("templates\CostingSheetTemplate.xlsx") # create new workbook
ws = wb['L18']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		row += 1
		ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			if item["VENDOR"] == "Vendor":
				for col in range(1,11):
					ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			if item["VENDOR"] == "Vendor":
				for col in range(1,11):
					ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			if item["VENDOR"] == "Vendor":
				for col in range(1,11):
					ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			if item["VENDOR"] == "Vendor":
				for col in range(1,11):
					ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L19']
#ws = wb.activ # selcet the active worksheet

row=1
for option in sorted(options):
	#for options:
	for item in options[option]["OUTFITTING PARTS"]:
		row += 1
		ws.cell(row = row, column = 1).value = option
		ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
		ws.cell(row = row, column = 3).value = item["PART NUMBER"][1:-1]
		ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
		ws.cell(row = row, column = 5).value = item["UOM"]
		ws.cell(row = row, column = 6).value = item["PRICE"]
		ws.cell(row = row, column = 7).value = item["19 QTY"]
	#for options:
	row += 1
	for item in options[option]["CANVAS PARTS"]:
		row += 1
		ws.cell(row = row, column = 1).value = option
		ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
		ws.cell(row = row, column = 3).value = item["PART NUMBER"][1:-1]
		ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
		ws.cell(row = row, column = 5).value = item["UOM"]
		ws.cell(row = row, column = 6).value = item["PRICE"]
		ws.cell(row = row, column = 7).value = item["19 QTY"]
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")