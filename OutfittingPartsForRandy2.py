import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
    options =  pickle.load(file)

wb = load_workbook("templates\CostingSheetTemplate.xlsx") # create new workbook
ws = wb['L18']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["18 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")
wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L19']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["19 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["19 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["19 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["19 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L20']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["20 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["20 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["20 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["20 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L21']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["21 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["21 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["21 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["21 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L22']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["22 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["22 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["22 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["22 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L23']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["23 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["23 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["23 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["23 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L24']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["24 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["24 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["24 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["24 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L25']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["25 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["25 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["25 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["25 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L26']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["26 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["26 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["26 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["26 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L27']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["27 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["27 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["27 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["27 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L28']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["28 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["28 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["28 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["28 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L29']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["29 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["29 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["29 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["29 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L30']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["30 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["30 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["30 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["30 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L31']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["31 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["31 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["31 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["31 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L32']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["32 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["32 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["32 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["32 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L33']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["33 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["33 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["33 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["33 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L34']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["34 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["34 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["34 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["34 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L35']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["35 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["35 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["35 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["35 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")

wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L36']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " outfitting/canvas/uph"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		if len(options[option]["OUTFITTING NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
			ws.cell(row = row, column = 1).font = blue
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["36 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) == 0 and len(options[option]["CANVAS PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " canvas"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["36 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["PAINT PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " paint"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["36 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
		ws.cell(row = row, column = 1).value = option + " fabrication"
		ws.cell(row = row, column = 1).font = bold
		ws.cell(row = row, column = 2).value =  options[option]["OPTION NAME"]
		ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
			row += 1
			ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
			ws.cell(row = row, column = 1).font = bold
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = item["VENDOR"]
			#if item["VENDOR"] == "Vendor":
				#for col in range(1,11):
					#ws.cell(row = row, column = col).font = red
			ws.cell(row = row, column = 2).value = item["VENDOR PART"]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = float(item["PRICE"])
			ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 5).value = item["UOM"]
			ws.cell(row = row, column = 6).value = item["36 QTY"]
			ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 8).value = 0
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
	if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["PAINT PARTS"]) + len(options[option]["FABRICATION PARTS"]) > 0:
		row += 1
			
wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")