import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
	options =  pickle.load(file)

kill=[]
for option in options:
	co = options.get(option+" - CO")
	cr = options.get(option+" - CR")
	if co:
		kill.append(option+" - CO")
		options[option]["FABRICATION PARTS"] += co["FABRICATION PARTS"]
		options[option]["PAINT PARTS"] += co["PAINT PARTS"]
	if cr:
		kill.append(option+" - CR")
		options[option]["FABRICATION PARTS"] += cr["FABRICATION PARTS"]
		options[option]["PAINT PARTS"] += cr["PAINT PARTS"]

for option in kill:
	del options[option]
		
wb = load_workbook("templates\CostingSheetTemplate.xlsx") # create new workbook
ws = wb['L18']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	
	if len(options[option]["OUTFITTING PARTS"]) + if len(options[option]["CANVAS"]) > 0:
		rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS"]
		for item in sorted(rigging, key=lambda k: (k["VENDOR"], k["PART NUMBER"])):
			pass
	
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		for item in sorted(options[option]["OUTFITTING PARTS"], key=lambda k: (k["VENDOR"], k["PART NUMBER"])):
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option + " Canvas"
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'

	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option + " Paint"
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option + " Fabrication"
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")
wb = load_workbook("output\OutfittingCostingSheetTemplateForRandy.xlsx") # create new workbook
ws = wb['L19']
bold = Font(bold=True, underline="single")
red = Font(color="FF0000")
blue = Font(color="0000FF")
#ws = wb.active # selcet the active worksheet
row = 1
for option in sorted(options):
	if len(options[option]["OUTFITTING PARTS"]) > 0:
		for item in options[option]["OUTFITTING PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			
	if len(options[option]["CANVAS PARTS"]) > 0:
		for item in options[option]["CANVAS PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option + " Canvas"
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'

	if len(options[option]["PAINT PARTS"]) > 0:
		for item in options[option]["PAINT PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option + " Paint"
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'
			
	if len(options[option]["FABRICATION PARTS"]) > 0:
		for item in options[option]["FABRICATION PARTS"]:
			row += 1
			print(option, item["PART NUMBER"])
			ws.cell(row = row, column = 1).value = option + " Fabrication"
			ws.cell(row = row, column = 2).value = item["VENDOR"]
			ws.cell(row = row, column = 3).value = item["VENDOR PART"]
			ws.cell(row = row, column = 4).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 5).value = float(item["PRICE"])
			ws.cell(row = row, column = 5).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 6).value = item["UOM"]
			ws.cell(row = row, column = 7).value = item["18 QTY"]
			ws.cell(row = row, column = 8).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
			ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 9).value = 0
			ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
			ws.cell(row = row, column = 10).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
			ws.cell(row = row, column = 10).number_format = '$#,##0.00;[Red]-$#,##0.00'

wb.save(r"output\OutfittingCostingSheetTemplateForRandy.xlsx")