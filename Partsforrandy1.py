import pickle
from openpyxl import load_workbook


# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
    options =  pickle.load(file)

wb = load_workbook("templates\CostingSheetTemplate.xlsx") # create new workbook
#ws = wb['L18']
#ws = wb.active # selcet the active worksheet

for row(skiprows = 1):
	ws = wb['L18']
		#for item in options[option]["OUTFITTING PARTS"]:
		for item in options[option]["OUTFITTING PARTS"]["PART NUMBER"]:
			#else if item["PART NUMBER"] = "":
			#else
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = item["PRICE"]
			ws.cell(row = row, column = 5).value = item["18 QTY"]
		#for item in options[option]["CANVAS PARTS"]:
		for item in options[option]["CANVAS PARTS"]["PART NUMBER"]:
			#else if item["PART NUMBER"] = "":
			#else
			ws.cell(row = row, column = 1).value = option
			ws.cell(row = row, column = 2).value = item["PART NUMBER"][1:-1]
			ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
			ws.cell(row = row, column = 4).value = item["PRICE"]
			ws.cell(row = row, column = 5).value = item["18 QTY"]
			
wb.save(r"output\CostingSheetTemplateForRandy.xlsx")