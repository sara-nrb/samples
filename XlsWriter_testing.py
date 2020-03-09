from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


wb = load_workbook("templates\\10121.xlsx") # create new workbook
ws = wb.active
ws.cell(row = 1, column = 1).value = 2
ws.cell(row = 1, column = 1).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
wb.save(r"output\dollars.xlsx")