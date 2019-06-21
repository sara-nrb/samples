#!/usr/bin/env python

import pickle
from openpyxl import load_workbook


# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
    options =  pickle.load(file)

wb = load_workbook("templates\eos_template.xlsx") # create new workbook
ws = wb.active # selcet the active worksheet

for row, option in enumerate(sorted(options), start = 2):
    value = options[option]["EOS DEPARTMENT"]
    if value is None:
        value = ""
    ws.cell(row = row, column = 1).value = option
    ws.cell(row = row, column = 2).value  = value[11:]  # skip "Category - "

wb.save(r"output\eos_from_template.xlsx")