#!/usr/bin/env python

from openpyxl import load_workbook
from excel import ExcelDocument
from pathlib import Path
from pythoncom import pywintypes
import os
import sys

"""
 __   __  ___       __        _______   _____  ___    __    _____  ___    _______   
|"  |/  \|  "|     /""\      /"      \ (\"   \|"  \  |" \  (\"   \|"  \  /" _   "|  
|'  /    \:  |    /    \    |:        ||.\\   \    | ||  | |.\\   \    |(: ( \___)  
|: /'        |   /' /\  \   |_____/   )|: \.   \\  | |:  | |: \.   \\  | \/ \       
 \//  /\'    |  //  __'  \   //      / |.  \    \. | |.  | |.  \    \. | //  \ ___  
 /   /  \\   | /   /  \\  \ |:  __   \ |    \    \ | /\  |\|    \    \ |(:   _(  _| 
|___/    \___|(___/    \___)|__|  \___) \___|\____\)(__\_|_)\___|\____\) \_______) 

   - This will modify the spreadseets in "K:\Links\2020\Options"
   - Make backups before testing the sheets or at the very least limit testing
   - limit testing with "break" at end of for loop so only executes once
   - by setting debud = True
   - replacements = [column-on-input-sheet, cell-on-option-sheet, None-replacement-value]
   - Example to replace "OPTION NAME", "OPTION NOTES", "ADVERTISED RETAIL"
   -   replacements = [["B", "B2", ""], ["C", "B3", ""], ["D", "B11", "0"]]
   - 
   
"""

debug = False
header_row = True
input_sheet = r"C:\Temp\Sample.xlsx"
path = r"K:\Links\2020\Options"
#path = r"C:\Temp\Options"
options_column = "A"
replacements = [["B", "B11", "0"]]

# load workbook, no need to close it is closed upon reading
wb = load_workbook(input_sheet) 
ws = wb.active # selcet the active worksheet

excel = ExcelDocument(False) # open Excel do not use existing copy

for row, cell in enumerate(ws[options_column], start =1):
    if row == 1 and header_row: # only needed if ROW 1 has titles
        continue
    file = os.path.join(path, str(cell.value) +".xlsx")
    try:
        excel.open(file, 3) # open file and update links
        excel.set_visible(False)
        excel.display_alerts(False)
        
        debug_output = "| {} | ".format(ws[options_column+str(row)].value )
        for column, cell, default in replacements:
            value = ws[column+str(row)].value
            if value is None or value == "None":
                value = default
            debug_output += "{} | ".format(str(value))
            excel.set_value(cell, value)

        print(debug_output)
        excel.save()
    except pywintypes.com_error as error:
        print("| Error: | {} |".format(str(error)))
    except:
        print("| Error: | {} |".format(str(sys.exc_info()[0])))

    if debug:
        break

excel.quit()
