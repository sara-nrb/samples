#!/usr/bin/env python

from openpyxl import load_workbook
from excel import ExcelDocument
from pathlib import Path
import os

r"""
 __   __  ___       __        _______   _____  ___    __    _____  ___    _______   
|"  |/  \|  "|     /""\      /"      \ (\"   \|"  \  |" \  (\"   \|"  \  /" _   "|  
|'  /    \:  |    /    \    |:        ||.\\   \    | ||  | |.\\   \    |(: ( \___)  
|: /'        |   /' /\  \   |_____/   )|: \.   \\  | |:  | |: \.   \\  | \/ \       
 \//  /\'    |  //  __'  \   //      / |.  \    \. | |.  | |.  \    \. | //  \ ___  
 /   /  \\   | /   /  \\  \ |:  __   \ |    \    \ | /\  |\|    \    \ |(:   _(  _| 
|___/    \___|(___/    \___)|__|  \___) \___|\____\)(__\_|_)\___|\____\) \_______) 

   - This will modify the spreadseets in "K:\Links\2020\Options"
   - Make backups before testing the sheets or at the very least limit testing
   - limit testing with "break" at end of for loop so only executes once
   - by setting debud = True

"""

debug = True
# load workbook, no need to close it is closed upon reading
wb = load_workbook("templates\\revised_pricing.xlsx") 
ws = wb.active # selcet the active worksheet
# A/col_1 = Option, B/col_2 = Old_Retial, C/col_3 = New_Retail
# will access with ws.cell(row = row, column = column).value

excel = ExcelDocument(False)

for row, cell in enumerate(ws['A'], start=1):
    if row == 1: continue # only needed if ROW 1 has titles

    file = os.path.join("K:\\Links\\2022\\Options\\" + str(cell.value) + ".xlsx")
    excel.open(file, 3) # open file and update links
    excel.set_visible(False)
    excel.display_alerts(False)
	
    # get all values from input sheet	
    new_retail = ws.cell(row = row, column = 3).value

    # write all values to output sheet
    excel.set_value("C8", new_retail)
	
    excel.save()
    excel.close()
    # for debug purposes and only running once
    if debug:
        print(ws.cell(row = row, column = 1).value, ws.cell(row = row, column = 2).value, ws.cell(row = row, column = 3).value)
        break

excel.quit()
