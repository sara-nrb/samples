import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime

# load data from pickle
with open(r"K:\Links\\New folder\new folder.pickle", "rb") as file:
    options =  pickle.load(file)

lengths = [slice('BOAT SIZES')] 
print (lengths)
 
wb = load_workbook(r"C:\Development\samples\templates\boatTemplate.xlsx") # create new workbook
ws = wb["COSTING SHEET"]

for option in sorted(options):
    bold = Font(bold=True)
    row = 1
    for length in sorted(lengths):

        if len (option['FABRICATION PARTS']) > 0:
                lengths = boats[boat][str('BOAT SIZES')]
                ws.cell(row = 3, column = 3).value = str(length) + " " + option
                ws.cell(row = 5, column = 3).value = "Fabrication Materials"
                ws.cell(row = 7, column = 1).value = "Vendor"
                ws.cell(row = 7, column = 1).font = bold 
                ws.cell(row = 7, column = 2).font = bold
                ws.cell(row = 7, column = 3).value = "Description"
                ws.cell(row = 7, column = 3).font = bold
                ws.cell(row = 7, column = 4).value = "Cost Per Pound"
                ws.cell(row = 7, column = 4).font = bold
                ws.cell(row = 7, column = 5).value = "UOM"
                ws.cell(row = 7, column = 5).font = bold
                ws.cell(row = 7, column = 6).value = "Pounds"
                ws.cell(row = 7, column = 6).font = bold
                ws.cell(row = 7, column = 7).value = "Sub Total"
                ws.cell(row = 7, column = 7).font = bold
                ws.cell(row = 7, column = 8).value = "Shipping"
                ws.cell(row = 7, column = 8).font = bold
                ws.cell(row = 7, column = 9).value = "Total"
                ws.cell(row = 7, column = 9).font = bold
        if len(boats[boat]["FABRICATION PARTS"]) > 0:
                row += 1
                ws.cell(row = row, column = 1).value = item["VENDOR"]
                ws.cell(row = row, column = 1).alignment = Alignment(horizontal='left')
                ws.cell(row = row, column = 2).value = item["VENDOR PART"]
                ws.cell(row = row, column = 2).alignment = Alignment(horizontal='left')
                ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
                ws.cell(row = row, column = 4).value = float(item["PRICE"])
                ws.cell(row = row, column = 4).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                ws.cell(row = row, column = 5).value = item["UOM"]
                ws.cell(row = row, column = 6).value = item[str(length) + " QTY"]
                ws.cell(row = row, column = 7).value = "=SUM(D" + str(row) + "*F" + str(row) + ")"
                ws.cell(row = row, column = 7).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                ws.cell(row = row, column = 8).value = 0
                ws.cell(row = row, column = 8).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
                ws.cell(row = row, column = 9).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
 
wb.save("C:\\Development\\samples\\output\\" + str(length) + " " + boat + "..xlsx")
