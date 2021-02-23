import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font



# load data from piclk
with open(r"K:\Links\2021\Options\New Folder\new folder.pickle", "rb") as file:
	options =  pickle.load(file)

lenghts = [
	"18",
	"20", 
	"21",
	"22",
	"23", 
	"24", 
	"25", 
	"27",
	"29",
	"31",
	"33",
	"35",
]

kill=[]
for option in options:
	co = options.get(option+" - CO")
	cr = options.get(option+" - CR")
	if co:
		kill.append(option+" - CO")
		options[option]["OUTFITTING PARTS"] += co["OUTFITTING PARTS"]
		options[option]["CANVAS PARTS"] += co["CANVAS PARTS"]
		options[option]["PAINT PARTS"] += co["PAINT PARTS"]
		options[option]["FABRICATION PARTS"] += co["FABRICATION PARTS"]
	if cr:
		kill.append(option+" - CR")
		options[option]["OUTFITTING PARTS"] += cr["OUTFITTING PARTS"]
		options[option]["CANVAS PARTS"] += cr["CANVAS PARTS"]
		options[option]["PAINT PARTS"] += cr["PAINT PARTS"]
		options[option]["FABRICATION PARTS"] += cr["FABRICATION PARTS"]

for option in kill:
	del options[option]

wb = load_workbook(r"templates\CostingSheetTemplate.xlsx") # create new workbook

for length in lenghts:
	ws = wb["L" + length]
	bold = Font(bold=True, underline="single", color="FF0000")
	red = Font(color="FF0000")
	blue = Font(color="0000FF")
	row = 1
	for option in sorted(options):

		if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["FABRICATION PARTS"]) + len(options[option]["PAINT PARTS"]) > 0:

				row += 1
				ws.cell(row = row, column = 1).value = option + " Outfitting"
				ws.cell(row = row, column = 1).font = bold 
				ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
				ws.cell(row = row, column = 2).font = bold
		if len(options[option]["OPTION NOTES"]) > 0:
				row += 1
				ws.cell(row = row, column = 1).value =  options[option]["OPTION NOTES"]
				ws.cell(row = row, column = 1).font = bold
		#if len(options[option]["OUTFITTING NOTES"]) > 0:
		#		row += 1
		#		ws.cell(row = row, column = 1).value =  options[option]["OUTFITTING NOTES"]
		#		ws.cell(row = row, column = 1).font = blue
		if len(options[option]["OUTFITTING PARTS"]) + len(options[option]["CANVAS PARTS"]) + len(options[option]["TRAILER PARTS"]) > 0:
			rigging = options[option]["OUTFITTING PARTS"] + options[option]["CANVAS PARTS"] + options[option]["TRAILER PARTS"]
			for item in rigging:
			
				row += 1
				print(option, item["PART NUMBER"])
				ws.cell(row = row, column = 1).value = item["VENDOR"]
				ws.cell(row = row, column = 2).value = item["VENDOR PART"]
				ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
				ws.cell(row = row, column = 4).value = float(item["PRICE"])
				ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 5).value = item["UOM"]
				ws.cell(row = row, column = 6).value = item[length + " QTY"]
				ws.cell(row = row, column = 7).value = "=SUM(F" + str(row) + "*D" + str(row) + ")"
				ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 8).value = 0
				ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
				ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
				
		if len(options[option]["FABRICATION PARTS"]) > 0:
				row += 1
				ws.cell(row = row, column = 1).value = option + " Fabrication"
				ws.cell(row = row, column = 1).font = bold 
				ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
				ws.cell(row = row, column = 2).font = bold
		if len(options[option]["FABRICATION PARTS"]) > 0:
			for item in options[option]["FABRICATION PARTS"]:
				row += 1
				print(option, item["PART NUMBER"])
				ws.cell(row = row, column = 1).value = item["VENDOR"]
				ws.cell(row = row, column = 2).value = item["VENDOR PART"]
				ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
				ws.cell(row = row, column = 4).value = float(item["PRICE"])
				ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 5).value = item["UOM"]
				ws.cell(row = row, column = 6).value = item[length + " QTY"]
				ws.cell(row = row, column = 7).value = "=SUM(F" + str(row) + "*D" + str(row) + ")"
				ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 8).value = 0
				ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
				ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
		
		if len(options[option]["PAINT PARTS"]) > 0:
				row += 1
				ws.cell(row = row, column = 1).value = option + " Paint"
				ws.cell(row = row, column = 1).font = bold 
				ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
				ws.cell(row = row, column = 2).font = bold
		if len(options[option]["PAINT PARTS"]) > 0:
			for item in options[option]["PAINT PARTS"]:
				row += 1
				print(option, item["PART NUMBER"])
				ws.cell(row = row, column = 1).value = item["VENDOR"]
				ws.cell(row = row, column = 2).value = item["VENDOR PART"]
				ws.cell(row = row, column = 3).value = item["DESCRIPTION"]
				ws.cell(row = row, column = 4).value = float(item["PRICE"])
				ws.cell(row = row, column = 4).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 5).value = item["UOM"]
				ws.cell(row = row, column = 6).value = item[length + " QTY"]
				ws.cell(row = row, column = 7).value = "=SUM(F" + str(row) + "*D" + str(row) + ")"
				ws.cell(row = row, column = 7).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 8).value = 0
				ws.cell(row = row, column = 8).number_format = '$#,##0.00;[Red]-$#,##0.00'
				ws.cell(row = row, column = 9).value = "=SUM(G" + str(row) + "+H" + str(row) + ")"
				ws.cell(row = row, column = 9).number_format = '$#,##0.00;[Red]-$#,##0.00'
		
wb.save(r"output\2021 Boat Options Parts Listing.xlsx")