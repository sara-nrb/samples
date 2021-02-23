from PyQt4 import QtGui # Import the PyQt4 module we'll need
from PyQt4 import QtCore
from PyQt4.QtCore import QSettings, QSize, QPoint
from PyQt4.QtCore import QThread, SIGNAL
from pathlib import Path
from dotenv import load_dotenv
from fields import topSection, startSections, endSections, partSection,  bottomSection
import openpyxl
import pickle
import sys # We need sys so that we can pass argv to QApplication
import os
import re
import MainWindow  # This file holds our MainWindow and all design related things


"""
Notes:
patch of the PyInstaller/depend/bindepend.py https://github.com/Loran425/pyinstaller/commit/14b6e65642e4b07a4358bab278019a48dedf7460

Lib\site-packages\PyQt4\pyuic4 MainWindow.ui  -o MainWindow.py

# Developed in C:\\Program Files\\NRB Pickle Options
Scripts\pyinstaller.exe --onefile --windowed --icon options.ico  --name "Options Fodler Pickler" "NRB Pickle Options FWW.spec" main.py

# Developed in C:\\Development\\nrb-pickle-options
Scripts\pyinstaller.exe --onefile --windowed --icon options.ico  --name "Options Fodler Pickler" "NRB Pickle Options Dev.spec" main.py

ToDo's
"""


class MainAppWindow(QtGui.QMainWindow, MainWindow.Ui_MainWindow):

    def __init__(self):
        super(self.__class__, self).__init__()
        self.setupUi(self)

        # set python environment
        if getattr(sys, 'frozen', False):
            bundle_dir = sys._MEIPASS
        else:
            # we are running in a normal Python environment
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        # load environmental variables
        load_dotenv(dotenv_path = Path(bundle_dir) / ".env")

        # set program icon
        self.setWindowIcon(QtGui.QIcon(os.path.join(bundle_dir, "pickle.ico")))

        # work in INI File Stuff here
        QtCore.QCoreApplication.setOrganizationName("NRB")
        QtCore.QCoreApplication.setOrganizationDomain("northriverboats.com")
        QtCore.QCoreApplication.setApplicationName("Options Fodler Pickler")
        self.settings = QSettings()
        

        # set variables
        self.background_thread = None
        self.exit_flag = False
        self.dir = self.settings.value("dir", os.getenv("DIR"))
        self.pickle_name = os.getenv("PICKLE")

        # set ui state
        self.actionCancel.setEnabled(False)
        self.btnCancel.hide()
        self.lePath.setText(self.dir)

        # set slots and signals
        self.actionExit.triggered.connect(self.closeEvent)
        self.actionAbout.triggered.connect(self.doAbout)
        self.btnBrowse.clicked.connect(self.browseEvent)
        self.btnRun.clicked.connect(self.startBackgroundTask)

    def doAbout(self, event):
        about_msg = "NRB Options Folder Pickler\n©2019 North River Boats\nBy Fred Warren"
        reply = QtGui.QMessageBox.information(self, 'About',
                         about_msg, QtGui.QMessageBox.Ok)

    def closeEvent(self, e):
        self._closeEvent(0)

    def _closeEvent(self, e):
        self.exit_flag = True
        self.settings.setValue("dir", self.dir)
        sys.exit(0)

    def browseEvent(self):
        default_dir = self.dir
        my_dir = QtGui.QFileDialog.getExistingDirectory(self, "Open a folder", default_dir, QtGui.QFileDialog.ShowDirsOnly)
        if my_dir != "":
            self.dir = my_dir
            self.update_progressbar(0)
        self.lePath.setText(self.dir)



    def startBackgroundTask(self):
        # hide / disable buttons and menu items as well as saving state
        self.block_actions()
        self.background_thread = background_thread(self.dir)

        self.connect(self.background_thread, SIGNAL('endBackgroundTask()'), self.endBackgroundTask)
        self.connect(self.background_thread, SIGNAL('update_statusbar(QString)'), self.update_statusbar)
        self.connect(self.background_thread, SIGNAL('update_label(QString)'), self.update_label)
        self.connect(self.background_thread, SIGNAL('update_progressbar(int)'), self.update_progressbar)
        self.btnCancel.clicked.connect(self.doAbort)    
        # Thread will self-terminate or be stopped via update_abort
        self.background_thread.start()

    def endBackgroundTask(self):
        self.unblock_actions()
        self.statusbar.showMessage("")

    def doAbort(self):
        self.background_thread.running = False
        self.unblock_actions

    def block_actions(self):        
        self.btnRun.setEnabled(False)
        self.btnCancel.setEnabled(True)
        self.btnRun.hide()
        self.btnCancel.show()
        self.actionRun.setEnabled(False)
        self.actionCancel.setEnabled(True)
        self.lePath.setReadOnly(True)
    
    def unblock_actions(self):
        self.btnRun.setEnabled(True)
        self.btnCancel.setEnabled(False)
        self.btnRun.show()
        self.btnCancel.hide()
        self.actionRun.setEnabled(True)
        self.actionCancel.setEnabled(False)
        self.lePath.setReadOnly(False)
        self.lblFile.setText("")

    def update_statusbar(self, message):
        self.statusbar.showMessage(message)

    def update_progressbar(self, num):
        self.progressBar.setValue(num)
        
    def update_label(self, label):
        self.lblFile.setText(label)


        
class background_thread(QThread):
    def __init__(self, dir):
        QThread.__init__(self)
        self.dir = dir

    def __del__(self):
        self.wait()

    def build_files_list(self, dir):
        self.emit(SIGNAL('update_statusbar(qString)'), "Finding files...")
        files = []
        for path in Path(dir).glob("[!~$]*.xlsx"):
            if not self.running:
                break
            files.append(path)
        self.emit(SIGNAL('update_statusbar(qString)'), "Found {} files to process".format(len(files)))
        return files

    def process_sheet(self, file):
        wb = openpyxl.load_workbook(file, data_only = True)
        ws = wb.active
        option = file.name[:-5]
        data = {}
        data["FILE"] = file.name
        data["FULLPATH"] = file.resolve()
        sections = ["FABRICATION", "CANVAS", "PAINT", "OUTFITTING"]
        
        # find where sections start
        starts = []
        for row in ws.iter_cols(min_col=1, max_col=1):
            for cell in row:
                if cell.value == "QTY":
                    starts.append(cell.row)

        # find where sections end
        ends = []
        for row in ws.iter_cols(min_col=5, max_col=5):
            for cell in row:
                if cell.value == "SUBTOTAL":
                    ends.append(cell.row)
        
        for name, column, row, default in topSection:
            value = ws.cell(column = column, row = row).value
            if value is None:
                value = default
            data[name] = value

        # Process top non-parts portion of sections
        for i, section in enumerate(sections):
            offset = starts[i]
            for name, column, row, default in startSections:
                value = ws.cell(column = column, row = row + offset).value
                if value is None:
                    value = default
                data[section + name] = value

        # Process bottom non-parts portion of sections
        for i, section in enumerate(sections):
            offset = ends[i]
            for name, column, row, default in endSections:
                value = ws.cell(column = column, row = row + offset).value
                if value is None:
                    value = default
                data[section + name] = value
       
        # Process parts portion of sections
        for i, section in enumerate(sections):
            data[section + " PARTS"] = []
            for offset in range(starts[i], ends[i]-1):
                if ws.cell(column = 2, row = 1 + offset).value is not None:
                    # print(option, ws.cell(column = 1, row = 1 + offset).value, ws.cell(column = 2, row = 1 + offset).value, ws.cell(column = 3, row = 1 + offset).value)
                    part = {}
                    for name, column, row, default in partSection:
                        value = ws.cell(column = column, row = row + offset).value
                        if value is None:
                            value = default
                        part[name] = value
                    data[section + " PARTS"].append(part)
        
        # Process bottom section
        offset = ends[3] + 5
        for name, column, row, default in bottomSection:
            value = ws.cell(column = column, row = row + offset).value
            if value is None:
                value = default
            data[name] = value

        # Handle Outfitting Notes
        if offset + 21 > ws.max_row:
            value = ""
        else:
            value = ws.cell(column = 1, row = offset + 21).value
        data["OUTFITTING NOTES"] = value
        
        return [str(data["OPTION NUMBER"]), data]

    def run(self):
        self.running = True
        self.emit(SIGNAL('update_progressbar(int)'), 0)
        files = self.build_files_list(self.dir)
        options = {}
        total_files = len(files)
        current_count = 0

        if not self.running:
            self.emit(SIGNAL('endBackgroundTask()'))
            return

        for file in files:
            if not self.running:
                break

            option, data = self.process_sheet(file)
            options[option] = data

            current_count += 1
            self.emit(SIGNAL('update_progressbar(int)'), int(float(current_count) / total_files * 100))
            self.emit(SIGNAL('update_label(QString)'), str(file))
            self.emit(SIGNAL('update_statusbar(QString)'), 'Pickeling %d of %d' % (current_count, total_files))

        # if we get to this point, pickle the results....
        file_name = os.path.join(self.dir, os.path.split(self.dir)[1].lower() + ".pickle")
        pickle.dump(options, open(file_name, 'wb'))
        self.emit(SIGNAL('endBackgroundTask()'))


def main():
    app = QtGui.QApplication(sys.argv)  # A new instance of QApplication
    form = MainAppWindow()              # We set the form to be our Main App Wehdiw (design)
    form.show()                         # Show the form
    app.exec_()                         # and execute the app


if __name__ == '__main__':              # if we're running file directly and not importing it
    main()                              # run the main function
                       # run the main function
