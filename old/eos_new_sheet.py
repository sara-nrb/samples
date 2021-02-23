#!/usr/bin/env python

import pickle
from openpyxl import Workbook
from openpyxl.styles import Font

# load data from piclk
with open(r"K:\Links\2020\Options\options.pickle", "rb") as file:
    options =  pickle.load(file)

wb = Workbook() # create new workbook
ws = wb.active # selcet the active worksheet

ws.column_dimensions["A"].width = 13 # set column width
ws["A1"] = "OPTIONS"
ws["A1"].font = Font(bold = True)

ws.column_dimensions["B"].width = 20
ws["B1"] = "DEPARTMENT"
ws["B1"].font = Font(bold = True)

for row, option in enumerate(sorted(options), start = 2):
    value = options[option]["EOS DEPARTMENT"]
    if value is None:
        value = ""
    ws.cell(row = row, column = 1).value = option
    ws.cell(row = row, column = 2).value  = value[11:]  # skip "Category - "

wb.save(r"output\eos_new_sheet.xlsx")