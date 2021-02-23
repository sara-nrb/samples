import pickle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime

# load data from piclk
with open(r"K:\Links\2021\Options\options.pickle", "rb") as file:
	options =  pickle.load(file)
    
#with open(r"K:\Links\2020\Yamaha Rigging\yamaha rigging.pickle", "rb") as file:
#	options =  pickle.load(file)

wb = load_workbook(r"templates\OVERHEAD.xlsx") # create new workbook
ws = wb["2021"]
row = 1

for option in sorted(options):
   # SHHT = options[option]["SHHT"]
    #SSOB = options[option]["SSOB"]
   # if SHHT == "Y" or if SSOB == "Y"
    bold = Font(bold=True, underline="single", color="FF0000")
    red = Font(color="FF0000")
    blue = Font(color="0000FF")
    

    row += 1
    ws.cell(row = row, column = 1).value = options[option]["OPTION NUMBER"]
    ws.cell(row = row, column = 2).value = options[option]["OPTION NAME"]
    ws.cell(row = row, column = 3).value = options[option]["CALCULATED RETAIL"]
    ws.cell(row = row, column = 3).number_format = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    ws.cell(row = row, column = 4).value = options[option]["TRAILER/MOTOR OVERHEAD"]
#   ws.cell(row = row, column = 9).value = options[option]["SSOB"]
#   ws.cell(row = row, column = 10).value = options[option]["LSOB"]
#   ws.cell(row = row, column = 11).value = options[option]["SHHT"]
#   ws.cell(row = row, column = 13).value = options[option]["23OS"]
#   ws.cell(row = row, column = 14).value = options[option]["SO"]
#   ws.cell(row = row, column = 15).value = options[option]["WXL"]
#   ws.cell(row = row, column = 16).value = options[option]["25OS"]
#   ws.cell(row = row, column = 17).value = options[option]["27OS"]
#   ws.cell(row = row, column = 18).value = options[option]["29OS"]
#   ws.cell(row = row, column = 19).value = options[option]["31OS"]
#   ws.cell(row = row, column = 20).value = options[option]["33OS"]
#   ws.cell(row = row, column = 21).value = options[option]["35OS"]
#   ws.cell(row = row, column = 22).value = options[option]["WASO"]
#   ws.cell(row = row, column = 23).value = options[option]["DV"]
#   ws.cell(row = row, column = 24).value = options[option]["C"]
#   ws.cell(row = row, column = 25).value = options[option]["OSP"]
#   ws.cell(row = row, column = 26).value = options[option]["S"]
    
wb.save(r"output\OVERHEAD.xlsx")