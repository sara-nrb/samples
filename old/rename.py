#!/usr/bin/env python

"""
 __   __  ___       __        _______   _____  ___    __    _____  ___    _______   
|"  |/  \|  "|     /""\      /"      \ (\"   \|"  \  |" \  (\"   \|"  \  /" _   "|  
|'  /    \:  |    /    \    |:        ||.\\   \    | ||  | |.\\   \    |(: ( \___)  
|: /'        |   /' /\  \   |_____/   )|: \.   \\  | |:  | |: \.   \\  | \/ \       
 \//  /\'    |  //  __'  \   //      / |.  \    \. | |.  | |.  \    \. | //  \ ___  
 /   /  \\   | /   /  \\  \ |:  __   \ |    \    \ | /\  |\|    \    \ |(:   _(  _| 
|___/    \___|(___/    \___)|__|  \___) \___|\____\)(__\_|_)\___|\____\) \_______) 

   - This will modify the spreadseets in "K:\Links\2020\Options"
   - Make backups before testing the sheets or at the very least limit testing
   - limit testing with "break" at end of for loop so only executes once
   - by setting debud = True

"""

import os
from openpyxl import load_workbook
illegal = ['\\', '/', ':', '*', '"', '<','>','|', '?']

path = "K:\\Links\\2020\\FAC Options\with Notes"
xls_files = [os.path.join(root, name)
  for root, dirs, files in os.walk(path)
  for name in files
  if name.startswith('FAC') and name.endswith(".xlsx")]

for file in xls_files:
  wb = load_workbook(file)
  ws = wb.active
  description = ws.cell(row=2, column=3).value
  for i in illegal:
    description = description.replace(i, '')
  new_name = file[:-5] + " " + description.strip() + ".xlsx"
  print("Renaming: {} to {}".format(file, new_name))
  os.rename(file, new_name)
